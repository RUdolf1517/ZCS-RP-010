"""
Панель управления достижениями учеников ZCS-RP-010
Student Achievement Management Panel

Загружает конфигурацию из переменных окружения (.env файл)
"""
import os
import io
import math
import logging
from datetime import datetime
from flask import Flask, redirect, render_template, request, session, url_for, send_file, flash
from sqlalchemy import delete, select, func
from sqlalchemy.orm import selectinload
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import json

# Загрузка переменных окружения из .env файла
# Это происходит ДО импорта других модулей, чтобы переменные были доступны везде
try:
    from dotenv import load_dotenv
    # Загружаем .env из текущей директории
    env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env')
    if os.path.exists(env_path):
        load_dotenv(env_path)
    else:
        # Пытаемся загрузить из текущей директории
        load_dotenv()
except ImportError:
    # python-dotenv не установлен, используем системные переменные окружения
    pass
except Exception as e:
    # Игнорируем ошибки загрузки .env, используем системные переменные
    pass

from database import (
    Student,
    AdminUserModel,
    Grade,
    SchoolClass,
    get_admin_user,
    get_db_session,
    init_db,
    search_students,
    authenticate_admin,
    create_default_admin,
    find_similar_students,
    create_database_backup,
    get_backup_list,
    restore_database_from_backup,
    get_all_users,
    create_admin_user,
    update_admin_user,
    delete_admin_user,
    check_user_permission,
    USER_ROLES,
    get_all_grades,
    get_classes_by_grade,
    create_grade,
    create_school_class,
    update_school_class,
)


def create_app():
    """
    Фабрика приложения Flask.
    Создает и настраивает приложение с использованием переменных окружения.
    """
    # Создаем приложение Flask
    app = Flask(__name__)

    # Настраиваем логирование
    app.logger.setLevel(logging.DEBUG)
    app.logger.info("Создание Flask приложения...")

    # Ключ для сессий и flash-сообщений (из переменных окружения)
    # Если не установлен, генерируем предупреждение
    secret_key = os.environ.get("FLASK_SECRET_KEY")
    if not secret_key:
        import warnings
        warnings.warn(
            "FLASK_SECRET_KEY не установлен! Используется небезопасный ключ по умолчанию. "
            "Установите FLASK_SECRET_KEY в .env файле для продакшена.",
            UserWarning
        )
        secret_key = "dev-secret-key-change-in-production"
    app.config["SECRET_KEY"] = secret_key

    # Создаем таблицы (если их еще нет)
    init_db()

    # Создаем дефолтного админа, если его нет
    create_default_admin()

    # JSON to object filter for jinja templates
    def safe_from_json(s):
        import json
        try:
            val = json.loads(s or '[]')
            if isinstance(val, list):
                return val
            return []
        except Exception:
            return []
    app.jinja_env.filters['from_json'] = safe_from_json

    # Главная: список карточек учеников (публичный просмотр)
    @app.route("/", methods=["GET"])
    def index():
        q = request.args.get("q")
        class_name = request.args.get("class")

        with next(get_db_session()) as db:
            # Загружаем учеников с классами и учителями заранее
            students_query = db.execute(
                select(Student).options(
                    selectinload(Student.school_class).selectinload(SchoolClass.class_teacher)
                ).order_by(Student.school_class_id, Student.full_name)
            ).scalars().all()

            # Применяем фильтры если заданы и группируем по классам
            classes_data = {}
            for student in students_query:
                # Применяем фильтры
                if q and q.lower() not in student.full_name.lower():
                    continue
                if class_name and class_name != student.class_name:
                    continue

                # Группируем по классам
                class_key = student.class_name
                if class_key not in classes_data:
                    classes_data[class_key] = {
                        'class_name': class_key,
                        'class_teacher': student.class_teacher,
                        'students': []
                    }
                classes_data[class_key]['students'].append(student)

            # Преобразуем в список и сортируем по названию класса
            classes_list = list(classes_data.values())
            classes_list.sort(key=lambda x: x['class_name'])

        return render_template("index.html", classes=classes_list, q=q or "", class_name=class_name or "")

    # Страница входа для админа
    @app.route("/admin/login", methods=["GET", "POST"])
    def admin_login():
        if request.method == "POST":
            username = request.form.get("username", "").strip()
            password = request.form.get("password", "").strip()

            app.logger.info(f"Попытка входа: пользователь '{username}'")

            # Аутентифицируем через новую систему
            admin = authenticate_admin(username, password)

            if admin:
                app.logger.info(f"Успешный вход: {admin.username} (роль: {admin.role})")
                session["admin_authenticated"] = True
                session["admin_username"] = admin.username
                session["admin_role"] = admin.role
                return redirect(url_for("admin_dashboard"))

            # Если данные неверные — показываем ошибку
            app.logger.warning(f"Неудачная попытка входа: {username}")
            return render_template("admin_login.html", error_message="Неверные учетные данные")

        app.logger.debug("Показ формы входа")
        return render_template("admin_login.html")

    # Выход из админ-панели
    @app.route("/admin/logout")
    def admin_logout():
        session.pop("admin_authenticated", None)
        return redirect(url_for("admin_login"))

    # Проверка аутентификации и ролей
    def require_admin():
        if not session.get("admin_authenticated"):
            return redirect(url_for("admin_login"))
        return True

    def require_admin_role(required_roles=None):
        """Требует указанные роли. Если roles=None, требует только аутентификации."""
        auth_result = require_admin()
        if auth_result != True:
            return auth_result

        if required_roles is not None:
            if isinstance(required_roles, str):
                required_roles = [required_roles]
            elif not isinstance(required_roles, list):
                required_roles = [required_roles]

            user_role = session.get("admin_role")
            if user_role not in required_roles:
                flash("Недостаточно прав доступа", "error")
                return redirect(url_for("admin_dashboard"))
        return True

    def can_export_class_reports(class_id=None):
        """Проверяет, может ли пользователь экспортировать отчеты класса."""
        user_role = session.get("admin_role")
        user_id = session.get("admin_id")

        # Админ может экспортировать всё
        if user_role == "admin":
            return True

        # Классный руководитель может экспортировать только свой класс
        if user_role == "class_teacher" and class_id:
            with next(get_db_session()) as db:
                school_class = db.get(SchoolClass, class_id)
                if school_class and school_class.class_teacher_id == user_id:
                    return True

        return False

    def require_permission(permission: str):
        """Проверяет наличие права доступа у текущего пользователя."""
        auth_result = require_admin()
        if auth_result != True:
            return auth_result

        user_role = session.get("admin_role", "")
        if not check_user_permission(user_role, permission):
            flash("Недостаточно прав доступа", "error")
            return redirect(url_for("admin_dashboard"))
        return True

    # Админ-панель: иерархический вид классов и параллелей
    @app.route("/admin", methods=["GET"])
    def admin_dashboard():
        app.logger.info("Запрос к админ-панели")
        auth_result = require_admin()
        if auth_result != True:
            app.logger.warning("Доступ к админ-панели запрещен")
            return auth_result

        app.logger.debug("Загрузка данных админ-панели")
        with next(get_db_session()) as db:
            # Получаем все параллели с классами и учениками
            grades_query = db.execute(
                select(Grade).options(
                    selectinload(Grade.classes).selectinload(SchoolClass.students),
                    selectinload(Grade.classes).selectinload(SchoolClass.class_teacher)
                ).order_by(Grade.grade_number)
            ).scalars().all()

            # Преобразуем в обычные структуры данных для шаблона
            grades = []
            for grade in grades_query:
                grade_data = {
                    'id': grade.id,
                    'grade_name': grade.grade_name,
                    'grade_number': grade.grade_number,
                    'classes': []
                }

                total_students_in_grade = 0
                for school_class in grade.classes:
                    class_data = {
                        'id': school_class.id,
                        'class_name': school_class.class_name,
                        'students': len(school_class.students),
                        'class_teacher': {
                            'username': school_class.class_teacher.username if school_class.class_teacher else None,
                            'role': school_class.class_teacher.role if school_class.class_teacher else None
                        } if school_class.class_teacher else None,
                        'recent_students': [
                            {
                                'full_name': student.full_name,
                                'achievements_count': len(student.achievements.split(',')) if student.achievements else 0
                            } for student in school_class.students[:3]
                        ] if school_class.students else []
                    }
                    grade_data['classes'].append(class_data)
                    total_students_in_grade += len(school_class.students)

                grade_data['total_students'] = total_students_in_grade
                grade_data['classes_count'] = len(grade_data['classes'])
                grades.append(grade_data)

            # Статистика
            total_students = db.execute(select(func.count()).select_from(Student)).scalar() or 0
            total_classes = db.execute(select(func.count()).select_from(SchoolClass)).scalar() or 0
            active_users = db.execute(
                select(func.count()).select_from(AdminUserModel).where(AdminUserModel.is_active == "1")
            ).scalar() or 0

        # Рендерим шаблон и передаем данные
        return render_template(
            "admin_dashboard.html",
            grades=grades,
            total_students=total_students,
            total_classes=total_classes,
            active_users=active_users,
            user_roles=USER_ROLES,
        )

    # Управление параллелями (админ)
    @app.route("/admin/grades", methods=["GET", "POST"])
    def admin_grades():
        auth_result = require_admin_role(['admin'])
        if auth_result != True:
            return auth_result

        with next(get_db_session()) as db:
            if request.method == "POST":
                action = request.form.get('action')

                if action == 'create_grade':
                    grade_number = request.form.get('grade_number', type=int)
                    if grade_number and 1 <= grade_number <= 12:
                        try:
                            # Проверяем, существует ли уже такая параллель
                            existing = db.execute(
                                select(Grade).where(Grade.grade_number == grade_number)
                            ).scalar_one_or_none()

                            if existing:
                                flash(f"Параллель {grade_number} классов уже существует!", "warning")
                            else:
                                grade = Grade(grade_number=grade_number)
                                db.add(grade)
                                db.commit()
                                flash(f"Параллель {grade_number} классов создана!", "success")

                        except Exception as e:
                            db.rollback()
                            flash(f"Ошибка при создании параллели: {e}", "error")
                    else:
                        flash("Неверный номер параллели!", "error")

                elif action == 'delete_grade':
                    grade_id = request.form.get('grade_id', type=int)
                    if grade_id:
                        try:
                            grade = db.get(Grade, grade_id)
                            if grade:
                                # Проверяем, есть ли классы в параллели
                                classes_count = db.execute(
                                    select(func.count()).select_from(SchoolClass).where(SchoolClass.grade_id == grade_id)
                                ).scalar() or 0

                                if classes_count > 0:
                                    flash(f"Нельзя удалить параллель с {classes_count} классами!", "warning")
                                else:
                                    db.delete(grade)
                                    db.commit()
                                    flash("Параллель удалена!", "success")
                            else:
                                flash("Параллель не найдена!", "error")
                        except Exception as e:
                            db.rollback()
                            flash(f"Ошибка при удалении параллели: {e}", "error")

                return redirect(url_for('admin_grades'))

            # GET запрос - показываем список параллелей
            grades_query = db.execute(
                select(Grade).options(
                    selectinload(Grade.classes).selectinload(SchoolClass.students)
                ).order_by(Grade.grade_number)
            ).scalars().all()

            # Преобразуем в обычные структуры данных для шаблона
            grades = []
            for grade in grades_query:
                grade_data = {
                    'id': grade.id,
                    'grade_name': grade.grade_name,
                    'grade_number': grade.grade_number,
                    'classes': []
                }

                total_students_in_grade = 0
                for school_class in grade.classes:
                    total_students_in_grade += len(school_class.students)

                grade_data['total_students'] = total_students_in_grade
                grade_data['classes_count'] = len(grade.classes)
                grades.append(grade_data)

            return render_template("admin_grades.html", grades=grades)

    # Создание карточки ученика (админ)
    @app.route("/admin/students/new", methods=["GET", "POST"])
    @app.route("/admin/students/new/<int:class_id>", methods=["GET", "POST"])
    def admin_student_new(class_id=None):
        auth_result = require_admin_role(['admin', 'class_teacher'])
        if auth_result != True:
            return auth_result

        with next(get_db_session()) as db:
            # Загружаем параллели с классами и учителями
            grades_query = db.execute(
                select(Grade).options(
                    selectinload(Grade.classes).selectinload(SchoolClass.class_teacher)
                ).order_by(Grade.grade_number)
            ).scalars().all()

            # Преобразуем в обычные структуры данных для шаблона
            grades = []
            for grade in grades_query:
                grade_data = {
                    'id': grade.id,
                    'grade_name': grade.grade_name,
                    'grade_number': grade.grade_number,
                    'classes': [
                        {
                            'id': school_class.id,
                            'class_name': school_class.class_name,
                            'class_teacher': {
                                'username': school_class.class_teacher.username if school_class.class_teacher else None
                            } if school_class.class_teacher else None
                        } for school_class in grade.classes
                    ]
                }
                grades.append(grade_data)

            selected_class = None
            if class_id:
                selected_class_query = db.execute(
                    select(SchoolClass).options(selectinload(SchoolClass.grade)).where(SchoolClass.id == class_id)
                ).scalar_one_or_none()
                if selected_class_query:
                    selected_class = {
                        'id': selected_class_query.id,
                        'class_name': selected_class_query.class_name
                    }

        if request.method == "POST":
            full_name = request.form.get("full_name", "").strip()
            school_class_id = request.form.get("school_class_id")

            ach_titles = request.form.getlist("ach_title[]")
            ach_levels = request.form.getlist("ach_level[]")
            ach_results = request.form.getlist("ach_result[]")
            ach_years = request.form.getlist("ach_year[]")
            ach_dates = request.form.getlist("ach_date[]")
            achievements = []
            for i in range(len(ach_titles)):
                if not (ach_titles[i].strip() and ach_levels[i] and ach_results[i] and ach_years[i] and ach_dates[i]):
                    continue
                achievements.append({
                    "title": ach_titles[i].strip(),
                    "level": ach_levels[i],
                    "result": ach_results[i],
                    "year": ach_years[i],
                    "date": ach_dates[i],
                })
            achievements_json = json.dumps(achievements, ensure_ascii=False)

            if not (full_name and school_class_id):
                return render_template("admin_student_form.html",
                                     error_message="Заполните все обязательные поля.",
                                     achievements=json.dumps(achievements, ensure_ascii=False),
                                     grades=grades,
                                     selected_class=selected_class)

            # Проверяем на возможное дублирование
            with next(get_db_session()) as db:
                school_class = db.get(SchoolClass, int(school_class_id))
                if not school_class:
                    return render_template("admin_student_form.html",
                                         error_message="Выбранный класс не найден.",
                                         achievements=json.dumps(achievements, ensure_ascii=False),
                                         grades=grades,
                                         selected_class=selected_class)

                # Ищем похожих учеников в том же классе
                similar_students = db.execute(
                    select(Student).where(
                        Student.school_class_id == int(school_class_id),
                        Student.full_name.ilike(f"%{full_name}%")
                    )
                ).scalars().all()

                if similar_students:
                    # Показываем предупреждение о дублировании
                    return render_template("admin_student_form.html",
                                         error_message="Возможное дублирование! Найдены похожие ученики в этом классе.",
                                         similar_students=similar_students,
                                         achievements=json.dumps(achievements, ensure_ascii=False),
                                         grades=grades,
                                         selected_class=selected_class,
                                         form_data={
                                             'full_name': full_name,
                                             'school_class_id': school_class_id
                                         })

                s = Student(
                    school_class_id=int(school_class_id),
                    full_name=full_name,
                    achievements=achievements_json or None,
                )
                db.add(s)
                db.commit()
            return redirect(url_for("admin_class_view", class_id=school_class_id))

        return render_template("admin_student_form.html",
                             achievements='[]',
                             grades=grades,
                             selected_class=selected_class)

    # Редактирование карточки ученика (админ)
    @app.route("/admin/students/<int:student_id>/edit", methods=["GET", "POST"])
    def admin_student_edit(student_id: int):
        auth_result = require_admin_role(['admin', 'class_teacher'])
        if auth_result != True:
            return auth_result

        with next(get_db_session()) as db:
            # Загружаем ученика с классом
            student = db.execute(
                select(Student).options(
                    selectinload(Student.school_class).selectinload(SchoolClass.grade)
                ).where(Student.id == student_id)
            ).scalar_one_or_none()

            if not student:
                flash("Ученик не найден", "error")
                return redirect(url_for("admin_dashboard"))

            # Загружаем параллели с классами и учителями
            grades_query = db.execute(
                select(Grade).options(
                    selectinload(Grade.classes).selectinload(SchoolClass.class_teacher)
                ).order_by(Grade.grade_number)
            ).scalars().all()

            # Преобразуем в обычные структуры данных для шаблона
            grades = []
            for grade in grades_query:
                grade_data = {
                    'id': grade.id,
                    'grade_name': grade.grade_name,
                    'grade_number': grade.grade_number,
                    'classes': [
                        {
                            'id': school_class.id,
                            'class_name': school_class.class_name,
                            'class_teacher': {
                                'username': school_class.class_teacher.username if school_class.class_teacher else None
                            } if school_class.class_teacher else None
                        } for school_class in grade.classes
                    ]
                }
                grades.append(grade_data)

            if request.method == "POST":
                full_name = request.form.get("full_name", "").strip()
                school_class_id = request.form.get("school_class_id")

                ach_titles = request.form.getlist("ach_title[]")
                ach_levels = request.form.getlist("ach_level[]")
                ach_results = request.form.getlist("ach_result[]")
                ach_years = request.form.getlist("ach_year[]")
                ach_dates = request.form.getlist("ach_date[]")
                achievements = []
                for i in range(len(ach_titles)):
                    if not (ach_titles[i].strip() and ach_levels[i] and ach_results[i] and ach_years[i] and ach_dates[i]):
                        continue
                    achievements.append({
                        "title": ach_titles[i].strip(),
                        "level": ach_levels[i],
                        "result": ach_results[i],
                        "year": ach_years[i],
                        "date": ach_dates[i],
                    })
                achievements_json = json.dumps(achievements, ensure_ascii=False)

                if not (full_name and school_class_id):
                    # Преобразуем ученика в обычную структуру для шаблона
                    student_data = {
                        'id': student.id,
                        'full_name': student.full_name,
                        'school_class_id': student.school_class_id,
                        'achievements': student.achievements
                    }
                    return render_template("admin_student_form.html",
                                         error_message="Заполните все обязательные поля.",
                                         achievements=json.dumps(achievements, ensure_ascii=False),
                                         student=student_data,
                                         grades=grades)

                school_class = db.get(SchoolClass, int(school_class_id))
                if not school_class:
                    # Преобразуем ученика в обычную структуру для шаблона
                    student_data = {
                        'id': student.id,
                        'full_name': student.full_name,
                        'school_class_id': student.school_class_id,
                        'achievements': student.achievements
                    }
                    return render_template("admin_student_form.html",
                                         error_message="Выбранный класс не найден.",
                                         achievements=json.dumps(achievements, ensure_ascii=False),
                                         student=student_data,
                                         grades=grades)

                student.full_name = full_name
                student.school_class_id = int(school_class_id)
                student.achievements = achievements_json or None
                db.commit()

                flash(f"Ученик {student.full_name} обновлен", "success")
                return redirect(url_for("admin_class_view", class_id=school_class_id))

        # Преобразуем ученика в обычную структуру для шаблона
        student_data = {
            'id': student.id,
            'full_name': student.full_name,
            'school_class_id': student.school_class_id,
            'achievements': student.achievements
        }

        achievements_json = student.achievements or '[]'
        return render_template("admin_student_form.html",
                             student=student_data,
                             achievements=safe_from_json(achievements_json),
                             grades=grades)

    # Удаление карточки ученика (админ)
    @app.route("/admin/students/<int:student_id>/delete", methods=["POST"])
    def admin_student_delete(student_id: int):
        auth_result = require_admin_role(['admin', 'class_teacher'])
        if auth_result != True:
            return auth_result

        with next(get_db_session()) as db:
            student = db.execute(
                select(Student).options(selectinload(Student.school_class)).where(Student.id == student_id)
            ).scalar_one_or_none()

            if student:
                # Проверяем права классного руководителя
                user_role = session.get("admin_role")
                user_id = session.get("admin_id")

                if user_role == "class_teacher" and student.school_class.class_teacher_id != user_id:
                    flash("Вы можете удалять только учеников своего класса", "error")
                    return redirect(url_for("admin_class_view", class_id=student.school_class_id))

                db.delete(student)
                db.commit()
        return redirect(url_for("admin_dashboard"))

    # Экспорт карточек учеников в Excel
    @app.route("/admin/export/excel")
    def admin_export_excel():
        auth_result = require_admin()
        if auth_result != True:
            return auth_result

        with next(get_db_session()) as db:
            # Загружаем учеников с классами и учителями для сортировки и экспорта
            students_query = db.execute(
                select(Student).options(
                    selectinload(Student.school_class).selectinload(SchoolClass.class_teacher)
                ).order_by(
                    Student.school_class_id.asc(), Student.full_name.asc()
                )
            ).scalars().all()

            # Сортируем по имени класса в Python (поскольку class_name - это property)
            students = sorted(students_query, key=lambda s: (s.class_name, s.full_name))

        # Создаем Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Карточки учеников"

        # Заголовки на русском
        headers = ["ID", "ФИО", "Класс", "Кл. руководитель", "Название конкурса", "Уровень", "Результат", "Год", "Дата участия", "Дата создания"]
        level_map = {'school':'Школьный','district':'Районный','region':'Региональный','russia':'Всероссийский','world':'Международный'}
        result_map = {'participant':'Участник','prize':'Призёр','winner':'Победитель'}
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Данные
        row = 2
        for student in students:
            parsed = []
            try:
                parsed = json.loads(student.achievements or "[]")
                if not isinstance(parsed, list):
                    parsed = []
            except Exception:
                parsed = []
            if parsed and isinstance(parsed, list) and "title" in parsed[0]:
                for ach in parsed:
                    ws.cell(row=row, column=1, value=student.id)
                    ws.cell(row=row, column=2, value=student.full_name)
                    ws.cell(row=row, column=3, value=student.class_name)
                    ws.cell(row=row, column=4, value=student.class_teacher)
                    ws.cell(row=row, column=5, value=ach.get("title", ""))
                    ws.cell(row=row, column=6, value=level_map.get(ach.get("level", ""), ach.get("level", "")))
                    ws.cell(row=row, column=7, value=result_map.get(ach.get("result", ""), ach.get("result", "")))
                    ws.cell(row=row, column=8, value=ach.get("year", ""))
                    ws.cell(row=row, column=9, value=ach.get("date", ""))
                    ws.cell(row=row, column=10, value=student.created_at.strftime("%Y-%m-%d %H:%M") if student.created_at else "")
                    row += 1
            else:
                ws.cell(row=row, column=1, value=student.id)
                ws.cell(row=row, column=2, value=student.full_name)
                ws.cell(row=row, column=3, value=student.class_name)
                ws.cell(row=row, column=4, value=student.class_teacher)
                ws.cell(row=row, column=5, value=student.achievements or "")
                for c in range(6,11):
                    ws.cell(row=row, column=c, value="")
                ws.cell(row=row, column=10, value=student.created_at.strftime("%Y-%m-%d %H:%M") if student.created_at else "")
                row += 1

        # Автоподбор ширины колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Оставляем авто-высоту строк (по умолчанию), т.к. wrap_text=True на ячейке достижений

        # Сохраняем в память
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Возвращаем файл для скачивания
        filename = f"ученики_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    # Экспорт по классам: каждый класс на отдельном листе
    @app.route("/admin/export/excel_by_class")
    def admin_export_excel_by_class():
        auth_result = require_admin_role(['admin', 'class_teacher'])
        if auth_result != True:
            return auth_result

        with next(get_db_session()) as db:
            # Загружаем учеников с классами и учителями для сортировки
            students_query = db.execute(
                select(Student).options(
                    selectinload(Student.school_class).selectinload(SchoolClass.class_teacher)
                ).order_by(
                    Student.school_class_id.asc(), Student.full_name.asc()
                )
            ).scalars().all()

            # Сортируем по имени класса в Python
            students = sorted(students_query, key=lambda s: (s.class_name, s.full_name))

        # Группируем по классам
        from collections import defaultdict
        class_to_students = defaultdict(list)
        for s in students:
            class_to_students[s.class_name].append(s)

        wb = Workbook()
        # Удаляем дефолтный лист, создадим свои
        default_ws = wb.active
        wb.remove(default_ws)

        headers = ["ID", "ФИО", "Класс", "Кл. руководитель", "Название конкурса", "Уровень", "Результат", "Год", "Дата участия", "Дата создания"]
        level_map = {'school':'Школьный','district':'Районный','region':'Региональный','russia':'Всероссийский','world':'Международный'}
        result_map = {'participant':'Участник','prize':'Призёр','winner':'Победитель'}

        for class_name, class_students in class_to_students.items():
            ws = wb.create_sheet(title=str(class_name)[:31])
            # Заголовки
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            # Данные
            row = 2
            for student in class_students:
                parsed = []
                try:
                    parsed = json.loads(student.achievements or "[]")
                    if not isinstance(parsed, list):
                        parsed = []
                except Exception:
                    parsed = []
                if parsed and isinstance(parsed, list) and "title" in parsed[0]:
                    for ach in parsed:
                        ws.cell(row=row, column=1, value=student.id)
                        ws.cell(row=row, column=2, value=student.full_name)
                        ws.cell(row=row, column=3, value=student.class_name)
                        ws.cell(row=row, column=4, value=student.class_teacher)
                        ws.cell(row=row, column=5, value=ach.get("title", ""))
                        ws.cell(row=row, column=6, value=level_map.get(ach.get("level", ""), ach.get("level", "")))
                        ws.cell(row=row, column=7, value=result_map.get(ach.get("result", ""), ach.get("result", "")))
                        ws.cell(row=row, column=8, value=ach.get("year", ""))
                        ws.cell(row=row, column=9, value=ach.get("date", ""))
                        ws.cell(row=row, column=10, value=student.created_at.strftime("%Y-%m-%d %H:%M") if student.created_at else "")
                        row += 1
                else:
                    ws.cell(row=row, column=1, value=student.id)
                    ws.cell(row=row, column=2, value=student.full_name)
                    ws.cell(row=row, column=3, value=student.class_name)
                    ws.cell(row=row, column=4, value=student.class_teacher)
                    ws.cell(row=row, column=5, value=student.achievements or "")
                    for c in range(6,11):
                        ws.cell(row=row, column=c, value="")
                    ws.cell(row=row, column=10, value=student.created_at.strftime("%Y-%m-%d %H:%M") if student.created_at else "")
                    row += 1

            # Автоподбор ширины
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"ученики_по_классам_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    # Экспорт Excel по выбранному классу (через ?class=7А)
    @app.route("/admin/export/excel_class")
    def admin_export_excel_class():
        auth_result = require_admin()
        if auth_result != True:
            return auth_result

        class_name = request.args.get("class", "").strip()
        if not class_name:
            return redirect(url_for("admin_dashboard"))

        with next(get_db_session()) as db:
            stmt = select(Student).where(Student.class_name == class_name).order_by(Student.full_name.asc())
            students = db.execute(stmt).scalars().all()

        wb = Workbook()
        ws = wb.active
        ws.title = (str(class_name) or "Класс")[:31]

        headers = ["ID", "ФИО", "Класс", "Кл. руководитель", "Название конкурса", "Уровень", "Результат", "Год", "Дата участия", "Дата создания"]
        level_map = {'school':'Школьный','district':'Районный','region':'Региональный','russia':'Всероссийский','world':'Международный'}
        result_map = {'participant':'Участник','prize':'Призёр','winner':'Победитель'}
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        row = 2
        for student in students:
            parsed = []
            try:
                parsed = json.loads(student.achievements or "[]")
                if not isinstance(parsed, list):
                    parsed = []
            except Exception:
                parsed = []
            if parsed and isinstance(parsed, list) and "title" in parsed[0]:
                for ach in parsed:
                    ws.cell(row=row, column=1, value=student.id)
                    ws.cell(row=row, column=2, value=student.full_name)
                    ws.cell(row=row, column=3, value=student.class_name)
                    ws.cell(row=row, column=4, value=student.class_teacher)
                    ws.cell(row=row, column=5, value=ach.get("title", ""))
                    ws.cell(row=row, column=6, value=level_map.get(ach.get("level", ""), ach.get("level", "")))
                    ws.cell(row=row, column=7, value=result_map.get(ach.get("result", ""), ach.get("result", "")))
                    ws.cell(row=row, column=8, value=ach.get("year", ""))
                    ws.cell(row=row, column=9, value=ach.get("date", ""))
                    ws.cell(row=row, column=10, value=student.created_at.strftime("%Y-%m-%d %H:%M") if student.created_at else "")
                    row += 1
            else:
                ws.cell(row=row, column=1, value=student.id)
                ws.cell(row=row, column=2, value=student.full_name)
                ws.cell(row=row, column=3, value=student.class_name)
                ws.cell(row=row, column=4, value=student.class_teacher)
                ws.cell(row=row, column=5, value=student.achievements or "")
                for c in range(6,11):
                    ws.cell(row=row, column=c, value="")
                ws.cell(row=row, column=10, value=student.created_at.strftime("%Y-%m-%d %H:%M") if student.created_at else "")
                row += 1

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"ученики_{class_name}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    # Массовый импорт учеников из Excel
    @app.route("/admin/import/excel", methods=["GET", "POST"])
    def admin_import_excel():
        auth_result = require_admin_role()
        if auth_result != True:
            return auth_result

        if request.method == "POST":
            if 'excel_file' not in request.files:
                flash("Файл не выбран", "error")
                return redirect(request.url)

            file = request.files['excel_file']
            if file.filename == '':
                flash("Файл не выбран", "error")
                return redirect(request.url)

            if not file.filename.endswith(('.xlsx', '.xls')):
                flash("Неверный формат файла. Используйте .xlsx или .xls", "error")
                return redirect(request.url)

            try:
                from openpyxl import load_workbook
                wb = load_workbook(file)
                ws = wb.active

                imported_count = 0
                errors = []

                with next(get_db_session()) as db:
                    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                        # Ожидаемая структура: ФИО, Класс, Кл. руководитель, Достижения (опционально)
                        if not row or not row[0] or not row[1] or not row[2]:
                            continue  # Пропускаем пустые строки

                        full_name = str(row[0]).strip()
                        class_name = str(row[1]).strip()
                        class_teacher = str(row[2]).strip()
                        achievements_text = str(row[3]).strip() if len(row) > 3 and row[3] else None

                        # Проверяем дублирование
                        existing = db.execute(
                            select(Student).where(
                                Student.full_name == full_name,
                                Student.class_name == class_name
                            )
                        ).scalar_one_or_none()

                        if existing:
                            errors.append(f"Строка {row_idx}: {full_name} ({class_name}) - уже существует")
                            continue

                        # Создаем достижения в формате JSON если есть текст
                        achievements_json = None
                        if achievements_text:
                            # Простой парсинг: разделяем по точкам с запятой
                            ach_list = []
                            for ach_text in achievements_text.split(';'):
                                ach_text = ach_text.strip()
                                if ach_text:
                                    ach_list.append({
                                        "title": ach_text,
                                        "level": "school",
                                        "result": "participant",
                                        "year": "25/26",
                                        "date": datetime.now().strftime("%Y-%m-%d")
                                    })
                            if ach_list:
                                achievements_json = json.dumps(ach_list, ensure_ascii=False)

                        # Создаем ученика
                        student = Student(
                            full_name=full_name,
                            class_name=class_name,
                            class_teacher=class_teacher,
                            achievements=achievements_json
                        )
                        db.add(student)
                        imported_count += 1

                    db.commit()

                flash(f"Успешно импортировано {imported_count} учеников", "success")
                if errors:
                    flash(f"Ошибки импорта: {'; '.join(errors[:5])}", "warning")  # Показываем первые 5 ошибок

            except Exception as e:
                flash(f"Ошибка при импорте: {str(e)}", "error")

            return redirect(url_for("admin_dashboard"))

        return render_template("admin_import.html")

    # Управление бэкапами базы данных
    @app.route("/admin/backups", methods=["GET", "POST"])
    def admin_backups():
        auth_result = require_admin_role()
        if auth_result != True:
            return auth_result

        if request.method == "POST":
            action = request.form.get("action")

            if action == "create":
                try:
                    backup_path = create_database_backup()
                    flash(f"Бэкап создан: {os.path.basename(backup_path)}", "success")
                except Exception as e:
                    flash(f"Ошибка при создании бэкапа: {str(e)}", "error")

            elif action == "restore" and "backup_file" in request.form:
                backup_file = request.form.get("backup_file")
                try:
                    backup_path = os.path.join("backups", backup_file)
                    restore_database_from_backup(backup_path)
                    flash(f"База данных восстановлена из: {backup_file}", "success")
                    # Переинициализируем соединение с БД
                    init_db()
                except Exception as e:
                    flash(f"Ошибка при восстановлении: {str(e)}", "error")

        # Получаем список бэкапов
        backups = get_backup_list()

        return render_template("admin_backups.html", backups=backups)

    # Скачивание бэкапа
    @app.route("/admin/backups/download/<filename>")
    def admin_backup_download(filename):
        auth_result = require_admin_role()
        if auth_result != True:
            return auth_result

        backup_path = os.path.join("backups", filename)
        if not os.path.exists(backup_path):
            flash("Файл бэкапа не найден", "error")
            return redirect(url_for("admin_backups"))

        return send_file(
            backup_path,
            as_attachment=True,
            download_name=filename,
            mimetype="application/octet-stream"
        )

    # Управление пользователями
    @app.route("/admin/users", methods=["GET"])
    def admin_users():
        auth_result = require_admin_role()  # Только админы могут управлять пользователями
        if auth_result != True:
            return auth_result

        with next(get_db_session()) as db:
            users = get_all_users(db)

        return render_template("admin_users.html", users=users, user_roles=USER_ROLES)

    # Создание нового пользователя
    @app.route("/admin/users/new", methods=["GET", "POST"])
    def admin_user_new():
        auth_result = require_admin_role()
        if auth_result != True:
            return auth_result

        if request.method == "POST":
            username = request.form.get("username", "").strip()
            password = request.form.get("password", "").strip()
            confirm_password = request.form.get("confirm_password", "").strip()
            role = request.form.get("role", "teacher")

            if not username or not password:
                flash("Заполните все обязательные поля", "error")
                return redirect(request.url)

            if password != confirm_password:
                flash("Пароли не совпадают", "error")
                return redirect(request.url)

            if len(password) < 6:
                flash("Пароль должен содержать минимум 6 символов", "error")
                return redirect(request.url)

            try:
                with next(get_db_session()) as db:
                    new_user = create_admin_user(db, username, password, role)
                    flash(f"Пользователь {new_user.username} успешно создан", "success")
                return redirect(url_for("admin_users"))
            except ValueError as e:
                flash(str(e), "error")
                return redirect(request.url)

        return render_template("admin_user_form.html", user_roles=USER_ROLES)

    # Редактирование пользователя
    @app.route("/admin/users/<int:user_id>/edit", methods=["GET", "POST"])
    def admin_user_edit(user_id: int):
        auth_result = require_admin_role()
        if auth_result != True:
            return auth_result

        with next(get_db_session()) as db:
            user = db.get(AdminUserModel, user_id)
            if not user:
                flash("Пользователь не найден", "error")
                return redirect(url_for("admin_users"))

            if request.method == "POST":
                username = request.form.get("username", "").strip()
                password = request.form.get("password", "").strip()
                confirm_password = request.form.get("confirm_password", "").strip()
                role = request.form.get("role", user.role)
                is_active = request.form.get("is_active", "1")

                if not username:
                    flash("Имя пользователя обязательно", "error")
                    return redirect(request.url)

                if password and password != confirm_password:
                    flash("Пароли не совпадают", "error")
                    return redirect(request.url)

                if password and len(password) < 6:
                    flash("Пароль должен содержать минимум 6 символов", "error")
                    return redirect(request.url)

                try:
                    update_data = {
                        "username": username,
                        "role": role,
                        "is_active": is_active
                    }
                    if password:
                        update_data["password"] = password

                    updated_user = update_admin_user(db, user_id, **update_data)
                    flash(f"Пользователь {updated_user.username} обновлен", "success")
                    return redirect(url_for("admin_users"))
                except ValueError as e:
                    flash(str(e), "error")
                    return redirect(request.url)

            return render_template("admin_user_form.html", user=user, user_roles=USER_ROLES)

    # Удаление пользователя
    @app.route("/admin/users/<int:user_id>/delete", methods=["POST"])
    def admin_user_delete(user_id: int):
        auth_result = require_admin_role()
        if auth_result != True:
            return auth_result

        try:
            with next(get_db_session()) as db:
                # Получаем пользователя перед удалением
                user = db.get(AdminUserModel, user_id)
                if user:
                    username = user.username
                    delete_admin_user(db, user_id)
                    flash(f"Пользователь {username} удален", "success")
                else:
                    flash("Пользователь не найден", "error")
        except ValueError as e:
            flash(str(e), "error")

        return redirect(url_for("admin_users"))

    # Управление классами и параллелями
    @app.route("/admin/classes", methods=["GET"])
    def admin_classes():
        auth_result = require_admin_role()
        if auth_result != True:
            return auth_result

        grade_id = request.args.get('grade_id', type=int)

        with next(get_db_session()) as db:
            if grade_id:
                # Показываем только выбранную параллель
                grade = db.execute(
                    select(Grade).options(
                        selectinload(Grade.classes).selectinload(SchoolClass.students),
                        selectinload(Grade.classes).selectinload(SchoolClass.class_teacher)
                    ).where(Grade.id == grade_id)
                ).scalar_one_or_none()

                if not grade:
                    flash("Параллель не найдена", "error")
                    return redirect(url_for('admin_grades'))

                grades = [grade]
            else:
                # Показываем все параллели
                grades = db.execute(
                    select(Grade).options(
                        selectinload(Grade.classes).selectinload(SchoolClass.students),
                        selectinload(Grade.classes).selectinload(SchoolClass.class_teacher)
                    ).order_by(Grade.grade_number)
                ).scalars().all()

            # Преобразуем в обычные структуры данных для шаблона
            grades_data = []
            for grade in grades:
                grade_data = {
                    'id': grade.id,
                    'grade_name': grade.grade_name,
                    'grade_number': grade.grade_number,
                    'classes': []
                }

                for school_class in grade.classes:
                    class_data = {
                        'id': school_class.id,
                        'class_name': school_class.class_name,
                        'students_count': len(school_class.students),
                        'class_teacher': {
                            'username': school_class.class_teacher.username if school_class.class_teacher else None,
                            'role': school_class.class_teacher.role if school_class.class_teacher else None
                        } if school_class.class_teacher else None
                    }
                    grade_data['classes'].append(class_data)

                grade_data['classes_count'] = len(grade_data['classes'])
                grades_data.append(grade_data)

        return render_template("admin_classes.html", grades=grades_data, selected_grade_id=grade_id, user_roles=USER_ROLES)

    # Создание нового класса
    @app.route("/admin/classes/new", methods=["GET", "POST"])
    def admin_class_new():
        auth_result = require_admin_role()
        if auth_result != True:
            return auth_result

        with next(get_db_session()) as db:
            # Загружаем параллели с классами
            grades_query = db.execute(
                select(Grade).options(selectinload(Grade.classes)).order_by(Grade.grade_number)
            ).scalars().all()

            # Преобразуем в обычные структуры данных для шаблона
            grades = []
            for grade in grades_query:
                grade_data = {
                    'id': grade.id,
                    'grade_name': grade.grade_name,
                    'grade_number': grade.grade_number,
                    'classes': [
                        {
                            'id': school_class.id,
                            'class_name': school_class.class_name
                        } for school_class in grade.classes
                    ]
                }
                grades.append(grade_data)

            teachers = get_all_users(db)

        if request.method == "POST":
            grade_id = request.form.get("grade_id")
            class_letter = request.form.get("class_letter", "").strip().upper()
            class_teacher_id = request.form.get("class_teacher_id") or None
            if class_teacher_id:
                class_teacher_id = int(class_teacher_id)

            if not grade_id or not class_letter:
                flash("Заполните все обязательные поля", "error")
                return redirect(request.url)

            try:
                with next(get_db_session()) as db:
                    school_class = create_school_class(db, int(grade_id), class_letter, class_teacher_id)
                    app.logger.info(f"Класс {school_class.class_name} успешно создан администратором")
                    flash(f"Класс {school_class.class_name} успешно создан", "success")
                return redirect(url_for("admin_classes"))
            except ValueError as e:
                error_msg = str(e)
                app.logger.warning(f"Попытка создать дублирующий класс: {error_msg}")
                flash(error_msg, "error")
                return redirect(request.url)

        return render_template("admin_class_form.html", grades=grades, teachers=teachers, user_roles=USER_ROLES)

    # Редактирование класса
    @app.route("/admin/classes/<int:class_id>/edit", methods=["GET", "POST"])
    def admin_class_edit(class_id: int):
        auth_result = require_admin_role()
        if auth_result != True:
            return auth_result

        with next(get_db_session()) as db:
            school_class = db.get(SchoolClass, class_id)
            if not school_class:
                flash("Класс не найден", "error")
                return redirect(url_for("admin_classes"))

            grades = get_all_grades(db)
            teachers = get_all_users(db)

            if request.method == "POST":
                class_teacher_id = request.form.get("class_teacher_id") or None
                if class_teacher_id:
                    class_teacher_id = int(class_teacher_id)

                try:
                    updated_class = update_school_class(db, class_id, class_teacher_id)
                    flash(f"Класс {updated_class.class_name} обновлен", "success")
                    return redirect(url_for("admin_classes"))
                except ValueError as e:
                    flash(str(e), "error")
                    return redirect(request.url)

            return render_template("admin_class_form.html", school_class=school_class, grades=grades, teachers=teachers, user_roles=USER_ROLES)

    # Удаление класса
    @app.route("/admin/classes/<int:class_id>/delete", methods=["POST"])
    def admin_class_delete(class_id: int):
        auth_result = require_admin_role()
        if auth_result != True:
            return auth_result

        with next(get_db_session()) as db:
            school_class = db.get(SchoolClass, class_id)
            if not school_class:
                flash("Класс не найден", "error")
                return redirect(url_for("admin_classes"))

            # Проверяем, есть ли ученики в классе
            students_count = db.execute(
                select(func.count()).select_from(Student).where(Student.school_class_id == class_id)
            ).scalar() or 0

            if students_count > 0:
                flash(f"Нельзя удалить класс с {students_count} учениками. Сначала удалите всех учеников.", "error")
                return redirect(url_for("admin_class_view", class_id=class_id))

            try:
                class_name = school_class.class_name
                db.delete(school_class)
                db.commit()
                flash(f"Класс {class_name} удален", "success")
            except Exception as e:
                db.rollback()
                flash(f"Ошибка при удалении класса: {e}", "error")

        return redirect(url_for("admin_classes"))

    # Просмотр класса и учеников
    @app.route("/admin/classes/<int:class_id>", methods=["GET"])
    def admin_class_view(class_id: int):
        auth_result = require_admin()
        if auth_result != True:
            return auth_result

        with next(get_db_session()) as db:
            # Загружаем класс со связанными данными
            school_class = db.execute(
                select(SchoolClass).options(
                    selectinload(SchoolClass.grade),
                    selectinload(SchoolClass.class_teacher)
                ).where(SchoolClass.id == class_id)
            ).scalar_one_or_none()

            if not school_class:
                flash("Класс не найден", "error")
                return redirect(url_for("admin_classes"))

            students = search_students(db, school_class_id=class_id)

            # Преобразуем в обычные структуры данных для шаблона
            class_data = {
                'id': school_class.id,
                'class_name': school_class.class_name,
                'grade': {
                    'grade_name': school_class.grade.grade_name if school_class.grade else 'Не указана'
                },
                'class_teacher': {
                    'username': school_class.class_teacher.username if school_class.class_teacher else None,
                    'role': school_class.class_teacher.role if school_class.class_teacher else None
                } if school_class.class_teacher else None
            }

            return render_template("admin_class_view.html", school_class=class_data, students=students, user_roles=USER_ROLES)

    # Экспорт отчета класса (для классных руководителей)
    @app.route("/admin/classes/<int:class_id>/export")
    def admin_export_class_report(class_id: int):
        auth_result = require_admin()
        if auth_result != True:
            return auth_result

        # Проверяем права на экспорт этого класса
        if not can_export_class_reports(class_id):
            flash("У вас нет прав на экспорт отчета этого класса", "error")
            return redirect(url_for("admin_class_view", class_id=class_id))

        with next(get_db_session()) as db:
            # Загружаем класс с учениками
            school_class = db.execute(
                select(SchoolClass).options(selectinload(SchoolClass.students)).where(SchoolClass.id == class_id)
            ).scalar_one_or_none()

            if not school_class:
                flash("Класс не найден", "error")
                return redirect(url_for("admin_classes"))

            # Создаем Excel файл для этого класса
            wb = Workbook()
            ws = wb.active
            ws.title = f"Отчет класса {school_class.class_name}"

            # Заголовки
            headers = ["ID", "ФИО", "Название конкурса", "Уровень", "Результат", "Год", "Дата участия", "Дата создания"]
            level_map = {'school':'Школьный','district':'Районный','region':'Региональный','russia':'Всероссийский','world':'Международный'}
            result_map = {'participant':'Участник','prize':'Призёр','winner':'Победитель'}

            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num, value=header)

            row = 2
            for student in school_class.students:
                parsed = []
                try:
                    parsed = json.loads(student.achievements or "[]")
                    if not isinstance(parsed, list):
                        parsed = []
                except Exception:
                    parsed = []

                if parsed and isinstance(parsed, list) and "title" in parsed[0]:
                    for ach in parsed:
                        ws.cell(row=row, column=1, value=student.id)
                        ws.cell(row=row, column=2, value=student.full_name)
                        ws.cell(row=row, column=3, value=ach.get('title', ''))
                        ws.cell(row=row, column=4, value=level_map.get(ach.get('level', ''), ach.get('level', '')))
                        ws.cell(row=row, column=5, value=result_map.get(ach.get('result', ''), ach.get('result', '')))
                        ws.cell(row=row, column=6, value=ach.get('year', ''))
                        ws.cell(row=row, column=7, value=ach.get('date', ''))
                        ws.cell(row=row, column=8, value=student.created_at.strftime('%Y-%m-%d') if student.created_at else '')
                        row += 1
                else:
                    # Студент без достижений
                    ws.cell(row=row, column=1, value=student.id)
                    ws.cell(row=row, column=2, value=student.full_name)
                    ws.cell(row=row, column=3, value="Достижения не добавлены")
                    ws.cell(row=row, column=8, value=student.created_at.strftime('%Y-%m-%d') if student.created_at else '')
                    row += 1

            # Автонастройка ширины колонок
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Максимальная ширина 50
                ws.column_dimensions[column_letter].width = adjusted_width

            # Создаем response
            from io import BytesIO
            bio = BytesIO()
            wb.save(bio)
            bio.seek(0)

            response = send_file(
                bio,
                as_attachment=True,
                download_name=f'отчет_класса_{school_class.class_name}_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx',
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

            return response

    # Временный маршрут для проверки
    @app.route("/test")
    def test_route():
        return "Приложение работает! Доступ по HTTP работает."

    return app


# Создаем приложение
app = create_app()


if __name__ == "__main__":
    """
    Запуск приложения в режиме разработки.
    В продакшене используйте systemd service или WSGI сервер (gunicorn, uwsgi).
    """
    # Настраиваем логирование для более подробного вывода
    logging.basicConfig(
        level=logging.DEBUG,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.StreamHandler(),  # Вывод в консоль
            logging.FileHandler('app.log')  # Сохранение в файл
        ]
    )

    # Настраиваем логирование Werkzeug (HTTP запросы)
    werkzeug_logger = logging.getLogger('werkzeug')
    werkzeug_logger.setLevel(logging.INFO)

    # Добавляем собственное логирование HTTP запросов
    @app.before_request
    def log_request_info():
        app.logger.info(f'{request.method} {request.url} - {request.remote_addr}')

    @app.after_request
    def log_response_info(response):
        app.logger.info(f'Response: {response.status_code}')
        return response

    # Включаем подробные логи для нашего приложения
    app_logger = logging.getLogger('app')
    app_logger.setLevel(logging.DEBUG)

    # Загружаем настройки из переменных окружения
    host = os.environ.get("FLASK_HOST", "0.0.0.0")
    port = int(os.environ.get("FLASK_PORT", "5000"))
    debug = os.environ.get("FLASK_DEBUG", "True").lower() == "true"

    print(f"Запуск панели управления учениками ZCS-RP-010")
    print(f"Адрес: http://{host}:{port}")
    print(f"Режим отладки: {'включен' if debug else 'выключен'}")
    print(f"Подробное логирование: ВКЛЮЧЕНО (уровень DEBUG)")
    print(f"Логи сохраняются в: app.log")
    print(f"Для остановки нажмите Ctrl+C")
    print()

    app.run(host=host, port=port, debug=debug)


